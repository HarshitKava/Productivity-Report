from msilib.schema import Font
from operator import index
from tokenize import group
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages,auth
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from .models import *
from ProductivityReport.models import *
from .forms import *#SiteEngDayForm,UpdateForm,CreateUserForm,Area_Input,AddCont,Add_Labour,Add_Lab_To_Contractor,ResetPasswordForm
from datetime import date
from .decorators import allowed_users, unauthenticated_user
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment
from datetime import datetime, timedelta

def Navbar(request):
    return render(request,'LabourReport/Navbar.html')

@unauthenticated_user
def LoginPage(request):
    if request.method == 'POST':
        userid = request.POST['userid']
        password = request.POST['pass']
        user_auth = authenticate(username=userid, password=password)
        admin = Group.objects.get(name="Admin").user_set.all()
        SE = Group.objects.get(name="Site Engineer").user_set.all()
        SLI = Group.objects.get(name="Site Labour Incharge").user_set.all()
        CLI = Group.objects.get(name="Camp Labour Incharge").user_set.all()
        Mang = Group.objects.get(name="Management").user_set.all()
        if user_auth is not None:
            if user_auth in SE:
                login(request,user_auth)
                return redirect('HomeSE')
            elif user_auth in admin:
                login(request,user_auth)
                return redirect('HomeAdmin')
            elif user_auth in SLI:
                login(request,user_auth)
                print("in")
                return redirect('HomeSLI')
            elif user_auth in CLI:
                login(request,user_auth)
                return redirect('HomeCLI')
            elif user_auth in Mang:
                login(request,user_auth)
                return redirect('HomeMang')
        else:
            messages.info(request,'Credential is incorrect')
            return redirect('/')
    return render(request,'LabourReport/LoginPage.html')

def LogoutUser(request):
    logout(request)
    return redirect('Login')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def HomeSE(request):
    return render(request,'LabourReport/HomeSE.html')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def AddDaySE(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    # print("d",d1,d2)
    Report=SiteEngDay.objects.filter(created_at__range=[d1,d2],Areaname=Areaname_id)
    form=SiteEngDayForm()
    if request.method =='POST':
        form=SiteEngDayForm(request.POST)
        print(form.errors)
        if form.is_valid():
            form.save()
        return redirect('AddDaySE')
    return render(request,'LabourReport/SiteEngAddDayData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def ViewDaySE(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=SiteEngDay.objects.filter(Areaname=Areaname_id)
    return render(request,'LabourReport/SiteEngViewDayData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def ViewNightSE(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=SiteEngNight.objects.filter(Areaname=Areaname_id)
    return render(request,'LabourReport/SiteEngViewNightData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def AddNightSE(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=SiteEngNight.objects.filter(created_at__range=[d1,d2],Areaname=Areaname_id)
    form=SiteEngNightForm()
    if request.method =='POST':
        form=SiteEngNightForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('/AddNightSE/')
    return render(request,'LabourReport/SiteEngAddNightData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def DeleteDaySE(request,i):
    new=SiteEngDay.objects.get(id=i)
    new.delete()
    return redirect('/AddDaySE/')
    

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def DeleteNightSE(request,i):
    new=SiteEngNight.objects.get(id=i)
    new.delete()
    return redirect('/AddNightSE/')
    

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def HomeAdmin(request):
    return render(request,'LabourReport/Admin/HomeAd.html')


@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def AddUser(request):
    registration_form=CreateUserForm()
    Area_input=Area_Input()
    # descending order of groups in User table
    User_data=User.objects.all().order_by('username')

    print(1)
    if request.method == 'POST':
        registration_form=CreateUserForm(request.POST)
        Area_input=Area_Input(request.POST)
        print(2)
        print(request.POST.get('groups'))
        print(registration_form.errors)
        if registration_form.is_valid():
            print(3)
            user=registration_form.save()
            if request.POST.get('groups')== "1":
                grp = Group.objects.get(name='Site Engineer')
                # grp=list(grp)
                user.groups.add(grp)
            elif request.POST.get('groups')== "2":
                grp = Group.objects.get(name='Site Labour Incharge')
                user.groups.add(grp)
            elif request.POST.get('groups')== "3":
                grp = Group.objects.get(name='Admin')
                user.groups.add(grp)
            elif request.POST.get('groups')== "4":
                grp = Group.objects.get(name='Management')
                user.groups.add(grp)
            elif request.POST.get('groups')== "5":
                grp = Group.objects.get(name='Camp Labour Incharge')
                user.groups.add(grp)
            if Area_input.is_valid():
                Area_input.save()
                print(4)
                return redirect('AddUser')
    return render(request,'LabourReport/Admin/User.html',{'form1':registration_form,'form2':Area_input,'User_data':User_data})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def AddContractor(request):
    form=AddCont()
    if request.method == 'POST':
        form=AddCont(request.POST)
        if form.is_valid():
            form.save()
            return redirect('HomeAdmin')
    return render(request,'LabourReport/Admin/Contractor.html',{'form':form})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])    
def AddLabours(request):
    form=Add_Labour()
    if request.method == 'POST':
        form=Add_Labour(request.POST)
        if form.is_valid():
            form.save()
            return redirect('HomeAdmin')
    return render(request,'LabourReport/Admin/Labour.html',{'form':form})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def LaboursOfContractor(request):
    form=Add_Lab_To_Contractor()
    if request.method == 'POST':
        form=Add_Lab_To_Contractor(request.POST)
        
        if form.is_valid():
            form.save()
            return redirect('HomeAdmin')
    return render(request,'LabourReport/Admin/LabourToCont.html',{'form':form})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def ResetPassword (request):
    form=ResetPasswordForm()
    if request.method == 'POST':
        form=ResetPasswordForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('HomeAdmin')
    return render(request,'LabourReport/Admin/ResetPassword.html',{'form':form})

@login_required(login_url='Login')
# @allowed_users(allowed_roles=['Site Engineer'])
def load_labour(request):
    contractor_id = request.GET.get('contractor_id')
    print(contractor_id)
    contractor_name = ContractorDetail.objects.get(id=contractor_id)
    labourofCont = LabourOfContractor.objects.filter(ContractorName=contractor_name).order_by('LabourCategory')
    labour =AddLabour.objects.filter()
    print(labourofCont,labour)
    return render(request, 'LabourReport/Admin/labour_dropdown_list_options.html', {'labour': labour,'labourofCont':labourofCont})

@login_required(login_url='Login')
# @allowed_users(allowed_roles=['Site Engineer'])
def load_cat(request):
    Labour_id = request.GET.get('contractor_id')
    print(Labour_id)
    # CategoryOfDeployment_id = CategoryOfDeployment.objects.get(id=Labour_id)
    # print(CategoryOfDeployment_id)
    labour_name = LabourOfContractor.objects.get(id=Labour_id)
    print(labour_name,type(labour_name),str(labour_name))
    labour =AddLabour.objects.get(LabourCategory=labour_name)
    print(labour,type(labour))
    Category=CategoryOfDeployment.objects.filter(ActivityName=labour)
    print(Category,type(Category))
    # {'labour': labour,'labourofCont':labourofCont}
    return render(request, 'LabourReport/Admin/category_dropdown_list_options.html',{'Category':Category})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def HomeSLI(request):
    return render(request,'LabourReport/SLI/HomeSLI.html')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def AddDaySLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=SLIDay.objects.filter(created_at__range=[d1,d2])
    form=SLIDayForm()
    if request.method =='POST':
        form=SLIDayForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('AddDaySLI')
    return render(request,'LabourReport/SLI/SLIAddDayData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def DeleteDaySLI(request,i,A,N,C,L,H):
    today = date.today()
    Report=SLIDay.objects.all()
    new=SLIDay.objects.get(id=i)
    updateData={
        'ContName': N,
        'AreaName': A,
        'LaborCat': C,
        'NoLabor': L,
        'NoHelp': H,
    }
    form=SLIDayForm(initial=updateData)
    if request.method =='POST':
        new.delete()
        return redirect('AddDaySLI')
    return render(request,'LabourReport/SLI/SLIDelDayData.html',{'Report':Report,'form':form,'today':today,'new':new})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def AddNightSLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=SLINight.objects.filter(created_at__range=[d1,d2])
    form=SLINightForm()
    if request.method =='POST':
        form=SLINightForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('AddNightSLI')
    return render(request,'LabourReport/SLI/SLIAddNightData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def DeleteNightSLI(request,i,A,N,C,L,H):
    today = date.today()
    Report=SLINight.objects.all()
    new=SLINight.objects.get(id=i)
    updateData={
        'ContName': N,
        'AreaName': A,
        'LaborCat': C,
        'NoLabor': L,
        'NoHelp': H,
    }
    form=SLINightForm(initial=updateData)
    if request.method =='POST':
        new.delete()
        return redirect('AddNightSLI')
    return render(request,'LabourReport/SLI/SLIDelNightData.html',{'Report':Report,'form':form,'today':today,'new':new})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def ViewDaySLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=SLIDay.objects.filter()
    return render(request,'LabourReport/SLI/SLIViewDayData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def ViewNightSLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=SLINight.objects.filter()
    return render(request,'LabourReport/SLI/SLIViewNightData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def ViewDayCLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=CLIDay.objects.filter(Areaname=Areaname_id)
    return render(request,'LabourReport/CLI/CLIViewDayData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def ViewNightCLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=CLINight.objects.filter(Areaname=Areaname_id)
    return render(request,'LabourReport/CLI/CLIViewNightData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def HomeCLI(request):
    return render(request,'LabourReport/CLI/HomeCLI.html')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def AddDayCLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=CLIDay.objects.filter(created_at__range=[d1,d2])
    form=CLIDayForm()
    if request.method =='POST':
        form=CLIDayForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('AddDayCLI')
    return render(request,'LabourReport/CLI/CLIAddDayData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def DeleteDayCLI(request,i,A,N,C,L,H):
    today = date.today()
    Report=CLIDay.objects.all()
    new=CLIDay.objects.get(id=i)
    updateData={
        'ContName': N,
        'AreaName': A,
        'LaborCat': C,
        'NoLabor': L,
        'NoHelp': H,
    }
    form=CLIDayForm(initial=updateData)
    if request.method =='POST':
        new.delete()
        return redirect('AddDayCLI')
    return render(request,'LabourReport/CLI/CLIDelDayData.html',{'Report':Report,'form':form,'today':today,'new':new})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def AddNightCLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=CLINight.objects.filter(created_at__range=[d1,d2])
    form=CLINightForm()
    if request.method =='POST':
        form=CLINightForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('AddNightCLI')
    return render(request,'LabourReport/CLI/CLIAddNightData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def DeleteNightCLI(request,i,A,N,C,L,H):
    today = date.today()
    Report=CLINight.objects.all()
    new=CLINight.objects.get(id=i)
    updateData={
        'ContName': N,
        'AreaName': A,
        'LaborCat': C,
        'NoLabor': L,
        'NoHelp': H,
    }
    form=CLINightForm(initial=updateData)
    if request.method =='POST':
        new.delete()
        return redirect('AddNightCLI')
    return render(request,'LabourReport/CLI/CLIDelNightData.html',{'Report':Report,'form':form,'today':today,'new':new})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Management'])
def HomeMang(request):
    return render(request,'LabourReport/Management/HomeMang.html')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Management'])
def SiteReport(request):
    current_user = request.user
    if request.method == 'POST':
        shift = request.POST.get('shift')
        date = request.POST.get('date')
        date1=datetime.strptime(date, "%Y-%m-%d")
        tomorrow = date1 + timedelta(1)
        d1=date1.strftime("%Y-%m-%d")
        d2=tomorrow.strftime("%Y-%m-%d")
        # area = request.POST.get('area')
        # print(area)
        
        
        # print(shift,date,d1,d2,area,area_id)

        workbook = Workbook()
        # Get active worksheet/tab
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename="Deployment Report "+date+" "+shift+".xlsx"
        response['Content-Disposition'] = 'attachment; filename='+filename
        worksheet = workbook.active
        worksheet.title = 'Deployment Report'
        area_arr=["SBN","KV","DBM","RKP","MPZ","HBM","ALK","AIIMS","Casting Yard","Casting Yard QC","Casting Yard PM"]
        area_arr1=["Labour Category","","SBN","","KV","","DBM","","RKP","","MPZ","","HBM","","ALK","","AIIMS","","Casting Yard","","Casting Yard QC","","Casting Yard PM","","Total",""]
        area_arr2=["","","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI"]
        LabCat=[]

        # Deployment Sheets
        for col_num, column_title in enumerate(area_arr1[2:], 1):
                
            if column_title=="":
                    # merge with next cell
                worksheet.merge_cells(start_row=1, start_column=col_num-1+2, end_row=1, end_column=col_num+2)
                #     # align to center
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell = worksheet.cell(row=1, column=col_num+2)
                cell.value = column_title
            worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
            worksheet.cell(row=1, column=1).value = "Labour Category"
            worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        for col_num, column_title in enumerate(area_arr2[2:], 1):
            cell = worksheet.cell(row=2, column=col_num+2)
            cell.value = column_title
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for cell in worksheet["1:1"]:
            cell.font = Font(bold=True)
        for cell in worksheet["2:2"]:
            cell.font = Font(bold=True)
        
        row_num = 3
        for i in area_arr:
            area=i
            area_id=Area.objects.filter(AreaName=area)
            area_id=area_id[0].id
            if shift == 'Day':
                rows = SiteEngDay.objects.filter(Areaname=area_id,created_at__range=[d1,d2]).values_list('LabourCategory', 'NoLabor')
            elif shift == 'Night':
                rows = SiteEngNight.objects.filter(Areaname=area_id,created_at__range=[d1,d2]).values_list('LabourCategory', 'NoLabor')
            worksheet['A1'].font = Font(bold=True)
            print(rows)
            
            # row_num = row_num+ 1
            for row in rows:
                LabourCategory = LabourOfContractor.objects.filter(pk=row[0]).values_list('LabourCategory')
                LabourCategory =list(LabourCategory[0])
                LabourCategory = AddLabour.objects.filter(pk=LabourCategory[0]).values_list('LabourCategory')
                LabourCategory =list(LabourCategory[0])
                print(LabourCategory[0])
                # Define the data for each cell in the row 
                row = [
                    LabourCategory[0],
                    row[1],
                ]
                # Assign the data for each cell of the row 
                # index of area in area_arr
                ind=area_arr1.index(area)+1
                if LabourCategory[0] in LabCat:
                    print(1)
                    row_num1=LabCat.index(LabourCategory[0])+3
                    worksheet.cell(row=row_num1, column=ind).value = row[1]
                else:
                    print(2)
                    worksheet.cell(row=row_num, column=ind).value = row[1]
                    worksheet.cell(row=row_num, column=1).value = row[0]
                    worksheet.cell(row=row_num, column=1).font = Font(bold=True)
                    worksheet.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    worksheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
                    row_num = row_num + 1
                    LabCat.append(LabourCategory[0])
                    # Total of row
                    worksheet.cell(row=row_num-1, column=25).value = '=SUM(C'+str(row_num-1)+'+E'+str(row_num-1)+'+G'+str(row_num-1)+'+I'+str(row_num-1)+'+K'+str(row_num-1)+'+M'+str(row_num-1)+'+O'+str(row_num-1)+'+Q'+str(row_num-1)+'+S'+str(row_num-1)+'+U'+str(row_num-1)+'+W'+str(row_num-1)+')'
                    worksheet.cell(row=row_num-1, column=26).value = '=SUM(D'+str(row_num-1)+'+F'+str(row_num-1)+'+H'+str(row_num-1)+'+J'+str(row_num-1)+'+L'+str(row_num-1)+'+N'+str(row_num-1)+'+P'+str(row_num-1)+'+R'+str(row_num-1)+'+T'+str(row_num-1)+'+V'+str(row_num-1)+'+X'+str(row_num-1)+')'
        
        # Total of column
        worksheet.cell(row=row_num, column=1).value = 'Total'
        worksheet.cell(row=row_num, column=1).font = Font(bold=True)
        worksheet.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')
        worksheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
        for i in range(3,27):
            # ascii of C
            worksheet.cell(row=row_num, column=i).value = '=SUM('+chr(64+i)+'3:'+chr(64+i)+str(row_num-1)+')'
        # freeze C3
        worksheet.freeze_panes = 'C3'
            
        # Site Sheets
        for i in area_arr:
            area=i
            area_id=Area.objects.filter(AreaName=area)
            area_id=area_id[0].id
            # Create multiple sheets using openpyxl
            worksheet = workbook.create_sheet(area)
            # worksheet.title = area
            # worksheet = workbook.create_sheet(area)
            # xlfile = workbook.create_sheet(area)
            # count+=1
            # if count>1:
            #     xlfile = workbook.create_sheet(area)
            #     xlfile = workbook.active
            #     xlfile['A1'] = 'Date'

            # Define the titles for columns
            columns = ['Contractor Name', 'Labour Category', 'Category of Deployment','Structure', 'Deployment']
            row_num = 1
            # Assign the titles for each cell of the header
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
            if shift == 'Day':
                rows = SiteEngDay.objects.filter(Areaname=area_id,created_at__range=[d1,d2]).values_list('ContractorName', 'LabourCategory','CategoryName','StructureName', 'NoLabor')
            elif shift == 'Night':
                rows = SiteEngNight.objects.filter(Areaname=area_id,created_at__range=[d1,d2]).values_list('ContractorName', 'LabourCategory','CategoryName','StructureName', 'NoLabor')
            
            worksheet['A1'].font = Font(bold=True)
            for cell in worksheet["1:1"]:
                cell.font = Font(bold=True)
            tot_lab=tot_help=tot_tot=0
            rows=list(rows)
            for row in rows:
                row_num += 1
                ContractorName = ContractorDetail.objects.filter(pk=row[0]).values_list('ContractorName')
                ContName=list(ContractorName[0])
                LabourCategory = LabourOfContractor.objects.filter(pk=row[1]).values_list('LabourCategory')
                LabourCategory =list(LabourCategory[0])
                LabourCategory = AddLabour.objects.filter(pk=LabourCategory[0]).values_list('LabourCategory')
                LabourCategory =list(LabourCategory[0])
                CategoryName = CategoryOfDeployment.objects.filter(pk=row[2]).values_list('CategoryName')
                CategoryName =list(CategoryName[0])
                StructureName = Structure.objects.filter(pk=row[3]).values_list('StructureName')
                StructureName =list(StructureName[0])
                row=[
                    ContName[0],
                    LabourCategory[0],
                    CategoryName[0],
                    StructureName[0],
                    row[4],
                ]
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
            # write total formula in cell
            if len(rows)>0:
                worksheet['D'+str(row_num+1)] = 'Total'
                worksheet['E'+str(row_num+1)] = '=SUM(E2:E'+str(row_num)+')'
                worksheet['D'+str(row_num+1)].font = Font(bold=True)
            else:
                # merge cells
                worksheet.merge_cells('A'+str(row_num+1)+':E'+str(row_num+1))
                worksheet['A'+str(row_num+1)] = 'No Data Found'
                # align to center
                worksheet['A'+str(row_num+1)].alignment = Alignment(horizontal='center')
            
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try: # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length+1
                worksheet.column_dimensions[column].width = adjusted_width
        workbook.save(response)
        return response

    return render(request,'LabourReport/Management/SiteReport.html')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Management'])
def FinalReport(request):
    if request.method == 'POST':
        shift = request.POST.get('shift')
        date = request.POST.get('date')
        date1=datetime.strptime(date, "%Y-%m-%d")
        tomorrow = date1 + timedelta(1)
        d1=date1.strftime("%Y-%m-%d")
        d2=tomorrow.strftime("%Y-%m-%d")
        reporttype = request.POST.get('type')
        
        print(shift,date,d1,d2,reporttype)

        workbook = Workbook()
        
        
        if reporttype == 'Site':
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename="Site Report "+date+" "+shift+".xlsx"
            response['Content-Disposition'] = 'attachment; filename='+filename
            worksheet = workbook.active
            worksheet.title = 'Site Report'+date+' '+shift
            area=["SBN","KV","DBM","RKP","MPZ","HBM","ALK","AIIMS","Casting Yard","Casting Yard QC","Casting Yard PM"]

            columns = ['Contractor Name', 'Labour Category', 'Category of Deployment','Structure', 'Deployment']
            worksheet.column_dimensions['A'].width = len(columns[0])
            worksheet.column_dimensions['B'].width = len(columns[1])

            worksheet['A1'] = 'Contractor Name'
            worksheet['B1'] = 'Types of Labour'
            cureent_cell=worksheet.cell(row=1,column=ord('A')-64)
            cureent_cell.alignment = Alignment(horizontal='center', vertical='center')
            cureent_cell=worksheet.cell(row=1,column=ord('B')-64)
            cureent_cell.alignment = Alignment(horizontal='center', vertical='center')

        #     col="C"
        #     for i in area:
        #         worksheet.column_dimensions[col].width = len('Unskilled')
        #         worksheet.column_dimensions[chr(ord(col)+1)].width = len('Unskilled')
        #         worksheet[col+'1']= i
        #         cureent_cell=worksheet.cell(row=1,column=ord(col)-64)
        #         cureent_cell.alignment = Alignment(horizontal='center')
        #         worksheet[col+'2']= 'Skilled'
        #         cureent_cell=worksheet.cell(row=2,column=ord(col)-64)
        #         cureent_cell.alignment = Alignment(horizontal='center')
        #         worksheet[chr(ord(col)+1)+'2']= 'Unskilled'
        #         cureent_cell=worksheet.cell(row=2,column=ord(col)-64+1)
        #         cureent_cell.alignment = Alignment(horizontal='center')
        #         col=chr(ord(col)+2)
        #     worksheet.merge_cells('A1:A2')
        #     worksheet.merge_cells('B1:B2')
        #     worksheet.merge_cells('C1:D1')
        #     worksheet.merge_cells('E1:F1')
        #     worksheet.merge_cells('G1:H1')
        #     worksheet.merge_cells('I1:J1')
        #     worksheet.merge_cells('K1:L1')
        #     worksheet.merge_cells('M1:N1')
        #     worksheet.merge_cells('O1:P1')
        #     worksheet.merge_cells('Q1:R1')
        #     worksheet.merge_cells('S1:T1')
        #     worksheet.merge_cells('U1:V1')
        #     worksheet.merge_cells('W1:X1')
        #     #Bold
        #     for cell in worksheet["1:1"]:
        #         cell.font = Font(bold=True)
        #     for cell in worksheet["2:2"]:
        #         cell.font = Font(bold=True)
        #     row_num = 3
        #     Cont=[]
        #     Lab=[]
        #     for i in area:
        #         area_id=Area.objects.filter(AreaName=i)
        #         if area_id:
        #             area_id=area_id[0].id
                    
        #             if shift == 'Day':
        #                 rows = SiteEngDay.objects.filter(Areaname=area_id,created_at__range=[d1,d2]).values_list('ContractorName', 'LabourCategory','CategoryName','StructureName' 'NoLabor')
        #             elif shift == 'Night':
        #                 rows = SiteEngNight.objects.filter(Areaname=area_id,created_at__range=[d1,d2]).values_list('ContractorName', 'LabourCategory','CategoryName','StructureName' 'NoLabor')
                    
        #             rows=list(rows)
        #             print(rows)
        #             for row in rows:
        #                 ContractorName = ContractorDetail.objects.filter(pk=row[0]).values_list('ContractorName')
        #                 ContName=list(ContractorName[0])
                        
                        
                        
        #                 LabourCategory = LabourOfContractor.objects.filter(pk=row[1]).values_list('LabourCategory')
        #                 LabourCategory =list(LabourCategory[0])
        #                 LabourCategory = AddLabour.objects.filter(pk=LabourCategory[0]).values_list('LabourCategory')
        #                 LabourCategory =list(LabourCategory[0])
        #                 if (ContName[0] in Cont) and (LabourCategory[0] in Lab) and (Cont.index(ContName[0]) == Lab.index(LabourCategory[0])):
        #                     index=Cont.index(ContName[0])
        #                     ind=area.index(i)
        #                     worksheet[chr(ord('C')+ind*2)+str(3+int(index))]=row[2]
        #                     worksheet[chr(ord('C')+ind*2+1)+str(3+int(index))]=row[3]
        #                 else:
        #                     Cont.append(ContName[0])
        #                     worksheet['A'+str(row_num)]=ContName[0]
        #                     Lab.append(LabourCategory[0])
        #                     worksheet['B'+str(row_num)]=LabourCategory[0]
        #                     ind=area.index(i)
        #                     worksheet[chr(ord('C')+ind*2)+str(row_num)]=row[2]
        #                     worksheet[chr(ord('C')+ind*2+1)+str(row_num)]=row[3]
        #                     row_num += 1
        #     worksheet.freeze_panes='C3'
        # elif reporttype == 'Final':
        #     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        #     filename="Final Report "+date+" "+shift+".xlsx"
        #     response['Content-Disposition'] = 'attachment; filename='+filename
        #     worksheet = workbook.active
        #     worksheet.title = 'Final Report'+date+' '+shift
        #     area=["SBN","KV","DBM","RKP","MPZ","HBM","ALK","AIIMS","Casting Yard","Casting Yard QC","Casting Yard PM"]
        #     columns = ['Contractor Name', 'Types of Labour', 'Skilled', 'Unskilled' ]
        #     worksheet.column_dimensions['A'].width = len(columns[0])
        #     worksheet.column_dimensions['B'].width = len(columns[1])

        #     worksheet['A1'] = 'Contractor Name'
        #     worksheet['B1'] = 'Types of Labour'
        #     cureent_cell=worksheet.cell(row=1,column=ord('A')-64)
        #     cureent_cell.alignment = Alignment(horizontal='center', vertical='center')
        #     cureent_cell=worksheet.cell(row=1,column=ord('B')-64)
        #     cureent_cell.alignment = Alignment(horizontal='center', vertical='center')
        #     col="C"
        #     col_num=3
        #     for i in area:
        #         worksheet.column_dimensions[col].width = len('Unskilled')
        #         worksheet.column_dimensions[col].width = len('Unskilled')
        #         worksheet[col+'1']= i+"-SE"
        #         cureent_cell=worksheet.cell(row=1,column=col_num)
        #         cureent_cell.alignment = Alignment(horizontal='center')
        #         worksheet[col+'2']= 'Skilled'
        #         cureent_cell=worksheet.cell(row=2,column=col_num)
        #         cureent_cell.alignment = Alignment(horizontal='center')
        #         if col=='Z':
        #             col='AA'
        #             col_num=ord(col[1])-64+26
        #         elif len(col)==1:
        #             col=chr(ord(col)+1)
        #             col_num=ord(col)-64
        #         else:
        #             col=col[0]+chr(ord(col[1])+1)
        #             col_num=ord(col[1])-64+26
        #         worksheet[col+'2']= 'Unskilled'
                
        #         cureent_cell=worksheet.cell(row=2,column=col_num)
        #         cureent_cell.alignment = Alignment(horizontal='center')
        #         worksheet.column_dimensions[col].width = len('Unskilled')
        #         if col=='Z':
        #             col='AA'
        #             col_num=ord(col[1])-64+26
        #         elif len(col)==1:
        #             col=chr(ord(col)+1)
        #             col_num=ord(col)-64
        #         else:
        #             col=col[0]+chr(ord(col[1])+1)
        #             col_num=ord(col[1])-64+26

        #         worksheet.column_dimensions[col].width = len('Unskilled')
        #         worksheet[col+'1']= i+"-SLI"
        #         cureent_cell=worksheet.cell(row=1,column=col_num)
        #         cureent_cell.alignment = Alignment(horizontal='center')
        #         worksheet[col+'2']= 'Skilled'
        #         cureent_cell=worksheet.cell(row=2,column=col_num)
        #         cureent_cell.alignment = Alignment(horizontal='center')
        #         if col=='Z':
        #             col='AA'
        #             col_num=ord(col[1])-64+26
        #         elif len(col)==1:
        #             col=chr(ord(col)+1)
        #             col_num=ord(col)-64
        #         else:
        #             col=col[0]+chr(ord(col[1])+1)
        #             col_num=ord(col[1])-64+26
        #         worksheet[col+'2']= 'Unskilled'
        #         if col=='Z':
        #             col='AA'
        #             col_num=ord(col[1])-64+26
        #         elif len(col)==1:
        #             col=chr(ord(col)+1)
        #             col_num=ord(col)-64
        #         else:
        #             col=col[0]+chr(ord(col[1])+1)
        #             col_num=ord(col[1])-64+26
        #         cureent_cell=worksheet.cell(row=2,column=col_num)
        #         cureent_cell.alignment = Alignment(horizontal='center')
            
        #     cureent_cell=worksheet.cell(row=1,column=47)
        #     cureent_cell.alignment = Alignment(horizontal='center', vertical='center')
        #     cureent_cell=worksheet.cell(row=1,column=48)
        #     cureent_cell.alignment = Alignment(horizontal='center', vertical='center')
        #     cureent_cell=worksheet.cell(row=2,column=49)
        #     cureent_cell.alignment = Alignment(horizontal='center')
        #     cureent_cell=worksheet.cell(row=1,column=49)
        #     cureent_cell.alignment = Alignment(horizontal='center')
        #     cureent_cell=worksheet.cell(row=1,column=50)
        #     cureent_cell.alignment = Alignment(horizontal='center')
        #     cureent_cell=worksheet.cell(row=1,column=51)
        #     cureent_cell.alignment = Alignment(horizontal='center', vertical='center')
            
        #     worksheet['AU1']= 'Total-SE'
        #     worksheet['AV1']= 'Total-SLI'
        #     worksheet['AW1']= 'CLI'
        #     worksheet['AW2']= 'Skilled'
        #     worksheet['AX2']= 'Unskilled'
        #     worksheet['AY1']= 'Total'
        #     worksheet.merge_cells('A1:A2')
        #     worksheet.merge_cells('B1:B2')
        #     worksheet.merge_cells('C1:D1')
        #     worksheet.merge_cells('E1:F1')
        #     worksheet.merge_cells('G1:H1')
        #     worksheet.merge_cells('I1:J1')
        #     worksheet.merge_cells('K1:L1')
        #     worksheet.merge_cells('M1:N1')
        #     worksheet.merge_cells('O1:P1')
        #     worksheet.merge_cells('Q1:R1')
        #     worksheet.merge_cells('S1:T1')
        #     worksheet.merge_cells('U1:V1')
        #     worksheet.merge_cells('W1:X1')
        #     worksheet.merge_cells('Y1:Z1')
        #     worksheet.merge_cells('AA1:AB1')
        #     worksheet.merge_cells('AC1:AD1')
        #     worksheet.merge_cells('AE1:AF1')
        #     worksheet.merge_cells('AG1:AH1')
        #     worksheet.merge_cells('AI1:AJ1')
        #     worksheet.merge_cells('AK1:AL1')
        #     worksheet.merge_cells('AM1:AN1')
        #     worksheet.merge_cells('AO1:AP1')
        #     worksheet.merge_cells('AQ1:AR1')
        #     worksheet.merge_cells('AS1:AT1')
        #     worksheet.merge_cells('AU1:AU2')
        #     worksheet.merge_cells('AV1:AV2')
        #     worksheet.merge_cells('AW1:AX1')
        #     worksheet.merge_cells('AY1:AY2')
        #     for cell in worksheet["1:1"]:
        #         cell.font = Font(bold=True)
        #     for cell in worksheet["2:2"]:
        #         cell.font = Font(bold=True)
        #     row_num=3
        #     Cont=[]
        #     Lab=[]
        #     for i in area:
        #         area_id=Area.objects.filter(AreaName=i)
        #         if area_id:
        #             area_id=area_id[0].id
        #             if shift=='Day':
        #                 rows_SE = SiteEngDay.objects.filter(Areaname=area_id,created_at__range=[d1,d2]).values_list('ContractorName', 'LabourCategory', 'NoLabor', 'NoHelp')
        #                 rows_CLI =CLIDay.objects.filter(created_at__range=[d1,d2]).values_list('ContractorName', 'LabourCategory', 'NoLabor', 'NoHelp')
        #                 rows_SLI=SLIDay.objects.filter(Areaname=area_id,created_at__range=[d1,d2]).values_list('ContractorName', 'LabourCategory', 'NoLabor', 'NoHelp')
        #             elif shift=='Night':
        #                 rows_SE = SiteEngNight.objects.filter(Areaname=area_id,created_at__range=[d1,d2]).values_list('ContractorName', 'LabourCategory', 'NoLabor', 'NoHelp')
        #                 rows_CLI =CLINight.objects.filter(created_at__range=[d1,d2]).values_list('ContractorName', 'LabourCategory', 'NoLabor', 'NoHelp')
        #                 rows_SLI=SLINight.objects.filter(Areaname=area_id,created_at__range=[d1,d2]).values_list('ContractorName', 'LabourCategory', 'NoLabor', 'NoHelp')
                    
        #             row_SE=list(rows_SE)
        #             for row in row_SE:
        #                 ContractorName = ContractorDetail.objects.filter(pk=row[0]).values_list('ContractorName')
        #                 ContName=list(ContractorName[0])    
        #                 LabourCategory = LabourOfContractor.objects.filter(pk=row[1]).values_list('LabourCategory')
        #                 LabourCategory =list(LabourCategory[0])
        #                 LabourCategory = AddLabour.objects.filter(pk=LabourCategory[0]).values_list('LabourCategory')
        #                 LabourCategory =list(LabourCategory[0])
        #                 if (ContName[0] in Cont) and (LabourCategory[0] in Lab) and (Cont.index(ContName[0]) == Lab.index(LabourCategory[0])):
        #                     index=Cont.index(ContName[0])
        #                     ind=area.index(i)
        #                     if ind<6:
        #                         # print(ContName[0],chr(ord('C')+ind*4),str(3+int(index)))
        #                         worksheet[chr(ord('C')+ind*4)+str(3+int(index))]=row[2]
        #                         worksheet[chr(ord('C')+ind*4+1)+str(3+int(index))]=row[3]
        #                     else:
        #                         worksheet["A"+chr(ord('C')+ind*4-26)+str(3+int(index))]=row[2]
        #                         worksheet["A"+chr(ord('C')+ind*4+1-26)+str(3+int(index))]=row[3]
        #                 else:
        #                     Cont.append(ContName[0])
        #                     worksheet['A'+str(row_num)]=ContName[0]
        #                     Lab.append(LabourCategory[0])
        #                     worksheet['B'+str(row_num)]=LabourCategory[0]
        #                     ind=area.index(i)
        #                     if ind<6:
        #                         worksheet[chr(ord('C')+ind*4)+str(row_num)]=row[2]
        #                         worksheet[chr(ord('C')+ind*4+1)+str(row_num)]=row[3]
        #                         worksheet["AU"+str(row_num)]= "=SUM(C"+str(row_num)+"+D"+str(row_num)+"+G"+str(row_num)+"+H"+str(row_num)+"+K"+str(row_num)+"+L"+str(row_num)+"+O"+str(row_num)+"+P"+str(row_num)+"+S"+str(row_num)+"+T"+str(row_num)+"+W"+str(row_num)+"+X"+str(row_num)+"+AA"+str(row_num)+"+AB"+str(row_num)+"+AE"+str(row_num)+"+AF"+str(row_num)+"+AI"+str(row_num)+"+AJ"+str(row_num)+"+AM"+str(row_num)+"+AN"+str(row_num)+"+AQ"+str(row_num)+"+AR"+str(row_num)+")"
        #                     else:
        #                         worksheet["A"+chr(ord('C')+ind*4-26)+str(3+int(index))]=row[2]
        #                         worksheet["A"+chr(ord('C')+ind*4+1-26)+str(3+int(index))]=row[3]
        #                         worksheet["AU"+str(row_num)]= "=SUM(C"+str(row_num)+"+D"+str(row_num)+"+G"+str(row_num)+"+H"+str(row_num)+"+K"+str(row_num)+"+L"+str(row_num)+"+O"+str(row_num)+"+P"+str(row_num)+"+S"+str(row_num)+"+T"+str(row_num)+"+W"+str(row_num)+"+X"+str(row_num)+"+AA"+str(row_num)+"+AB"+str(row_num)+"+AE"+str(row_num)+"+AF"+str(row_num)+"+AI"+str(row_num)+"+AJ"+str(row_num)+"+AM"+str(row_num)+"+AN"+str(row_num)+"+AQ"+str(row_num)+"+AR"+str(row_num)+")"
        #                     row_num+=1
        #             for row in rows_SLI:
        #                 ContractorName = ContractorDetail.objects.filter(pk=row[0]).values_list('ContractorName')
        #                 ContName=list(ContractorName[0])    
        #                 LabourCategory = LabourOfContractor.objects.filter(pk=row[1]).values_list('LabourCategory')
        #                 LabourCategory =list(LabourCategory[0])
        #                 LabourCategory = AddLabour.objects.filter(pk=LabourCategory[0]).values_list('LabourCategory')
        #                 LabourCategory =list(LabourCategory[0])
        #                 if (ContName[0] in Cont) and (LabourCategory[0] in Lab) and (Cont.index(ContName[0]) == Lab.index(LabourCategory[0])):
        #                     index=Cont.index(ContName[0])
        #                     ind=area.index(i)
        #                     if ind<6:
        #                         # print("SLI-",ContName[0],chr(ord('E')+ind*4),str(3+int(index)))
        #                         worksheet[chr(ord('E')+ind*4)+str(3+int(index))]=row[2]
        #                         worksheet[chr(ord('E')+ind*4+1)+str(3+int(index))]=row[3]
        #                     else:
        #                         # print("SLI-",ContName[0],"A"+chr(ord('E')+ind*4-26),str(3+int(index)))
        #                         worksheet["A"+chr(ord('E')+ind*4-26)+str(3+int(index))]=row[2]
        #                         worksheet["A"+chr(ord('E')+ind*4+1-26)+str(3+int(index))]=row[3]
        #                 else:
        #                     Cont.append(ContName[0])
        #                     worksheet['A'+str(row_num)]=ContName[0]
        #                     Lab.append(LabourCategory[0])
        #                     worksheet['B'+str(row_num)]=LabourCategory[0]
        #                     ind=area.index(i)
        #                     if ind<6:
        #                         worksheet[chr(ord('E')+ind*4)+str(row_num)]=row[2]
        #                         worksheet[chr(ord('E')+ind*4+1)+str(row_num)]=row[3]
        #                         worksheet["AV"+str(row_num)]= "=SUM(E"+str(row_num)+"+F"+str(row_num)+"+I"+str(row_num)+"+J"+str(row_num)+"+M"+str(row_num)+"+N"+str(row_num)+"+Q"+str(row_num)+"+Q"+str(row_num)+"+R"+str(row_num)+"+U"+str(row_num)+"+V"+str(row_num)+"+Y"+str(row_num)+"+Z"+str(row_num)+"+AC"+str(row_num)+"+AD"+str(row_num)+"+AG"+str(row_num)+"+AH"+str(row_num)+"+AK"+str(row_num)+"+AL"+str(row_num)+"+AO"+str(row_num)+"+AP"+str(row_num)+"+AS"+str(row_num)+"+AT"+str(row_num)+")"
        #                     else:
        #                         worksheet["A"+chr(ord('E')+ind*4-26)+str(3+int(index))]=row[2]
        #                         worksheet["A"+chr(ord('E')+ind*4+1-26)+str(3+int(index))]=row[3]
        #                         worksheet["AV"+str(row_num)]= "=SUM(E"+str(row_num)+"+F"+str(row_num)+"+I"+str(row_num)+"+J"+str(row_num)+"+M"+str(row_num)+"+N"+str(row_num)+"+Q"+str(row_num)+"+Q"+str(row_num)+"+R"+str(row_num)+"+U"+str(row_num)+"+V"+str(row_num)+"+Y"+str(row_num)+"+Z"+str(row_num)+"+AC"+str(row_num)+"+AD"+str(row_num)+"+AG"+str(row_num)+"+AH"+str(row_num)+"+AK"+str(row_num)+"+AL"+str(row_num)+"+AO"+str(row_num)+"+AP"+str(row_num)+"+AS"+str(row_num)+"+AT"+str(row_num)+")"
        #                     row_num+=1
        #             for row in rows_CLI:
        #                 ContractorName = ContractorDetail.objects.filter(pk=row[0]).values_list('ContractorName')
        #                 ContName=list(ContractorName[0])    
        #                 LabourCategory = LabourOfContractor.objects.filter(pk=row[1]).values_list('LabourCategory')
        #                 LabourCategory =list(LabourCategory[0])
        #                 LabourCategory = AddLabour.objects.filter(pk=LabourCategory[0]).values_list('LabourCategory')
        #                 LabourCategory =list(LabourCategory[0])
        #                 if (ContName[0] in Cont) and (LabourCategory[0] in Lab) and (Cont.index(ContName[0]) == Lab.index(LabourCategory[0])):
        #                     index=Cont.index(ContName[0])
        #                     ind=area.index(i)
        #                     # print("SLI-",ContName[0],"A"+chr(ord('E')+ind*4-26),str(3+int(index)))
        #                     worksheet["AW"+str(3+int(index))]=row[2]
        #                     worksheet["AX"+str(3+int(index))]=row[3]
        #                 else:
        #                     Cont.append(ContName[0])
        #                     worksheet['A'+str(row_num)]=ContName[0]
        #                     Lab.append(LabourCategory[0])
        #                     worksheet['B'+str(row_num)]=LabourCategory[0]
        #                     worksheet["AW"+str(row_num)]=row[2]
        #                     worksheet["AX"+str(row_num)]=row[3]
        #                     worksheet["AY"+str(row_num)]= "=SUM(AW"+str(row_num)+"+AX"+str(row_num)+")"
        #                     row_num+=1
        # #freezing the first row
        # worksheet.freeze_panes = 'C3'

        workbook.save(response)
        # return response
    return render(request,'LabourReport/Management/FinalReport.html')